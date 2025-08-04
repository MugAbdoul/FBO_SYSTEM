from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models.admin import Admin, AdminRole, Gender
from app.models.applicant import Applicant
from app.models.organization_application import OrganizationApplication, ApplicationStatus
from app.models.supporting_document import SupportingDocument
from app.models.notification import Notification, NotificationType
from app.utils.auth import hash_password, admin_required, get_current_user
from app.utils.validators import validate_email, validate_phone, validate_password
from datetime import datetime, timedelta
from sqlalchemy import func, extract, and_, or_
import io
import csv
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.colors import black, blue
import openpyxl
from openpyxl.styles import Font, Alignment

bp = Blueprint('admin', __name__)

@bp.route('/users', methods=['GET'])
@admin_required(roles=[AdminRole.CEO, AdminRole.SECRETARY_GENERAL])
def get_users():
    """Get all admin users"""
    try:
        current_admin = get_current_user()
        
        # Build query based on current user's role
        query = Admin.query
        
        # Secretary General can only see users below their level
        if current_admin.role == AdminRole.SECRETARY_GENERAL:
            excluded_roles = [AdminRole.CEO, AdminRole.SECRETARY_GENERAL]
            query = query.filter(~Admin.role.in_(excluded_roles))
        
        users = query.order_by(Admin.created_at.desc()).all()
        
        return jsonify({
            'users': [user.to_dict() for user in users]
        })
        
    except Exception as e:
        return jsonify({'error': f'Failed to fetch users: {str(e)}'}), 500

@bp.route('/users', methods=['POST'])
@admin_required(roles=[AdminRole.CEO, AdminRole.SECRETARY_GENERAL])
def create_user():
    """Create a new admin user"""
    try:
        current_admin = get_current_user()
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['email', 'password', 'firstname', 'lastname', 
                          'phonenumber', 'role', 'gender']
        
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Validate email
        if not validate_email(data['email']):
            return jsonify({'error': 'Invalid email format'}), 400
        
        # Check if email already exists
        if Admin.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email already registered'}), 400
        
        # Validate password
        is_valid, message = validate_password(data['password'])
        if not is_valid:
            return jsonify({'error': message}), 400
        
        # Validate phone
        if not validate_phone(data['phonenumber']):
            return jsonify({'error': 'Invalid phone number format'}), 400
        
        # Validate role
        try:
            role = AdminRole(data['role'])
        except ValueError:
            return jsonify({'error': 'Invalid role'}), 400
        
        # Check role permissions
        if current_admin.role == AdminRole.SECRETARY_GENERAL:
            # Secretary General cannot create CEO or other Secretary General accounts
            forbidden_roles = [AdminRole.CEO, AdminRole.SECRETARY_GENERAL]
            if role in forbidden_roles:
                return jsonify({'error': 'Insufficient permissions to create this role'}), 403
        
        # Validate gender
        try:
            gender = Gender(data['gender'].upper())
        except ValueError:
            return jsonify({'error': 'Invalid gender'}), 400
        
        # Create new admin user
        admin = Admin(
            email=data['email'].lower(),
            password=hash_password(data['password']),
            firstname=data['firstname'].strip(),
            lastname=data['lastname'].strip(),
            phonenumber=data['phonenumber'].strip(),
            role=role,
            gender=gender,
            enabled=data.get('enabled', True)
        )
        
        db.session.add(admin)
        db.session.commit()
        
        # Create notification for the new user
        notification = Notification(
            admin_id=admin.id,
            type=NotificationType.STATUS_CHANGE,
            title='Welcome to RGB Portal',
            message=f'Your admin account has been created. You can now login with your credentials.'
        )
        db.session.add(notification)
        db.session.commit()
        
        return jsonify({
            'message': 'User created successfully',
            'user': admin.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create user: {str(e)}'}), 500

@bp.route('/users/<int:user_id>', methods=['PUT'])
@admin_required(roles=[AdminRole.CEO, AdminRole.SECRETARY_GENERAL])
def update_user(user_id):
    """Update an admin user"""
    try:
        current_admin = get_current_user()
        admin = Admin.query.get_or_404(user_id)
        data = request.get_json()
        
        # Check permissions
        if current_admin.role == AdminRole.SECRETARY_GENERAL:
            # Secretary General cannot edit CEO or other Secretary General accounts
            forbidden_roles = [AdminRole.CEO, AdminRole.SECRETARY_GENERAL]
            if admin.role in forbidden_roles:
                return jsonify({'error': 'Insufficient permissions to edit this user'}), 403
        
        # Validate email if provided
        if 'email' in data and data['email'] != admin.email:
            if not validate_email(data['email']):
                return jsonify({'error': 'Invalid email format'}), 400
            
            # Check if email already exists
            existing_admin = Admin.query.filter(
                and_(Admin.email == data['email'], Admin.id != user_id)
            ).first()
            if existing_admin:
                return jsonify({'error': 'Email already registered'}), 400
            
            admin.email = data['email'].lower()
        
        # Validate phone if provided
        if 'phonenumber' in data and data['phonenumber']:
            if not validate_phone(data['phonenumber']):
                return jsonify({'error': 'Invalid phone number format'}), 400
            admin.phonenumber = data['phonenumber'].strip()
        
        # Update basic fields
        updateable_fields = ['firstname', 'lastname']
        for field in updateable_fields:
            if field in data and data[field]:
                setattr(admin, field, data[field].strip())
        
        # Update role if provided and allowed
        if 'role' in data and data['role']:
            try:
                new_role = AdminRole(data['role'])
                
                # Check role permissions
                if current_admin.role == AdminRole.SECRETARY_GENERAL:
                    forbidden_roles = [AdminRole.CEO, AdminRole.SECRETARY_GENERAL]
                    if new_role in forbidden_roles:
                        return jsonify({'error': 'Insufficient permissions to assign this role'}), 403
                
                admin.role = new_role
            except ValueError:
                return jsonify({'error': 'Invalid role'}), 400
        
        # Update gender if provided
        if 'gender' in data and data['gender']:
            try:
                admin.gender = Gender(data['gender'].upper())
            except ValueError:
                return jsonify({'error': 'Invalid gender'}), 400
        
        # Update enabled status if provided
        if 'enabled' in data:
            admin.enabled = bool(data['enabled'])
        
        admin.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'message': 'User updated successfully',
            'user': admin.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to update user: {str(e)}'}), 500

@bp.route('/users/<int:user_id>', methods=['DELETE'])
@admin_required(roles=[AdminRole.CEO])
def delete_user(user_id):
    """Delete an admin user (CEO only)"""
    try:
        current_admin = get_current_user()
        admin = Admin.query.get_or_404(user_id)
        
        # Prevent self-deletion
        if admin.id == current_admin.id:
            return jsonify({'error': 'Cannot delete your own account'}), 400
        
        # Prevent deletion of other CEOs
        if admin.role == AdminRole.CEO:
            return jsonify({'error': 'Cannot delete other CEO accounts'}), 400
        
        # Check if user has processed applications
        processed_apps = OrganizationApplication.query.filter_by(processed_by_id=user_id).count()
        if processed_apps > 0:
            return jsonify({
                'error': f'Cannot delete user who has processed {processed_apps} applications. Consider disabling instead.'
            }), 400
        
        # Delete user's notifications first
        Notification.query.filter_by(admin_id=user_id).delete()
        
        # Delete the user
        db.session.delete(admin)
        db.session.commit()
        
        return jsonify({'message': 'User deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete user: {str(e)}'}), 500

@bp.route('/dashboard/stats', methods=['GET'])
@admin_required()
def get_dashboard_stats():
    """Get dashboard statistics"""
    try:
        current_admin = get_current_user()
        
        # Base statistics
        stats = {
            'total_applications': OrganizationApplication.query.count(),
            'total_users': Admin.query.filter_by(enabled=True).count(),
            'total_applicants': Applicant.query.filter_by(enabled=True).count(),
        }
        
        # Role-specific statistics
        if current_admin.role == AdminRole.CEO:
            # CEO gets comprehensive stats
            stats.update({
                'pending_review': OrganizationApplication.query.filter_by(status=ApplicationStatus.SG_REVIEW).count(),
                'approved': OrganizationApplication.query.filter_by(status=ApplicationStatus.APPROVED).count(),
                'rejected': OrganizationApplication.query.filter_by(status=ApplicationStatus.REJECTED).count(),
                'certificates_issued': OrganizationApplication.query.filter_by(status=ApplicationStatus.CERTIFICATE_ISSUED).count(),
            })
            
            # Applications by status
            status_counts = {}
            for status in ApplicationStatus:
                count = OrganizationApplication.query.filter_by(status=status).count()
                status_counts[status.value] = count
            stats['by_status'] = status_counts
            
            # Monthly statistics (last 12 months)
            monthly_stats = []
            for i in range(12):
                date = datetime.now() - timedelta(days=30 * i)
                start_date = date.replace(day=1)
                if i == 0:
                    end_date = datetime.now()
                else:
                    end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                
                count = OrganizationApplication.query.filter(
                    and_(
                        OrganizationApplication.submitted_at >= start_date,
                        OrganizationApplication.submitted_at <= end_date
                    )
                ).count()
                
                monthly_stats.append({
                    'month': date.month,
                    'year': date.year,
                    'count': count
                })
            
            stats['monthly'] = list(reversed(monthly_stats))
            
            # Processing time statistics
            avg_processing_time = db.session.query(
                func.avg(
                    func.extract('epoch', OrganizationApplication.last_modified) - 
                    func.extract('epoch', OrganizationApplication.submitted_at)
                ) / 86400  # Convert to days
            ).filter(
                OrganizationApplication.status.in_([ApplicationStatus.APPROVED, ApplicationStatus.REJECTED])
            ).scalar()
            
            stats['avg_processing_days'] = round(avg_processing_time, 1) if avg_processing_time else 0
            
        else:
            # Other roles get limited stats based on their responsibilities
            role_status_map = {
                AdminRole.FBO_OFFICER: [ApplicationStatus.PENDING, ApplicationStatus.UNDER_REVIEW, ApplicationStatus.REVIEWING_AGAIN],
                AdminRole.DIVISION_MANAGER: [ApplicationStatus.DRAFT],
                AdminRole.HOD: [ApplicationStatus.DM_REVIEW],
                AdminRole.SECRETARY_GENERAL: [ApplicationStatus.HOD_REVIEW]
            }
            
            if current_admin.role in role_status_map:
                stats['pending_review'] = OrganizationApplication.query.filter(
                    OrganizationApplication.status.in_(role_status_map[current_admin.role])
                ).count()
        
        # Recent activity
        recent_applications = OrganizationApplication.query.order_by(
            OrganizationApplication.submitted_at.desc()
        ).limit(5).all()
        
        stats['recent_applications'] = [
            {
                'id': app.id,
                'organization_name': app.organization_name,
                'status': app.status.value,
                'submitted_at': app.submitted_at.isoformat()
            } for app in recent_applications
        ]
        
        return jsonify({'stats': stats})
        
    except Exception as e:
        return jsonify({'error': f'Failed to get dashboard stats: {str(e)}'}), 500

@bp.route('/reports/generate', methods=['POST'])
@admin_required()
def generate_report():
    """Generate custom reports"""
    try:
        current_admin = get_current_user()
        data = request.get_json()
        
        # Validate required fields
        report_type = data.get('reportType', 'summary')
        start_date = data.get('startDate')
        end_date = data.get('endDate')
        status_filter = data.get('status', '')
        format_type = data.get('format', 'pdf')
        
        if not start_date or not end_date:
            return jsonify({'error': 'Start date and end date are required'}), 400
        
        # Parse dates
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)  # Include end date
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        
        # Build query
        query = OrganizationApplication.query.filter(
            and_(
                OrganizationApplication.submitted_at >= start_date,
                OrganizationApplication.submitted_at < end_date
            )
        )
        
        if status_filter:
            try:
                status = ApplicationStatus(status_filter)
                query = query.filter(OrganizationApplication.status == status)
            except ValueError:
                return jsonify({'error': 'Invalid status filter'}), 400
        
        applications = query.order_by(OrganizationApplication.submitted_at.desc()).all()
        
        # Generate report based on type and format
        if format_type == 'csv':
            return generate_csv_report(applications, report_type, start_date, end_date)
        elif format_type == 'excel':
            return generate_excel_report(applications, report_type, start_date, end_date)
        else:  # PDF
            return generate_pdf_report(applications, report_type, start_date, end_date, current_admin)
            
    except Exception as e:
        return jsonify({'error': f'Failed to generate report: {str(e)}'}), 500

def generate_csv_report(applications, report_type, start_date, end_date):
    """Generate CSV report"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    if report_type == 'detailed':
        headers = [
            'Application ID', 'Organization Name', 'Acronym', 'Applicant Name', 
            'Email', 'Phone', 'Address', 'Status', 'Submitted At',
            'Last Modified', 'Certificate Number', 'Cluster of Intervention'
        ]
    else:
        headers = [
            'Application ID', 'Organization Name', 'Applicant Name', 
            'Status', 'Submitted At'
        ]
    
    writer.writerow(headers)
    
    # Data
    for app in applications:
        if report_type == 'detailed':
            row = [
                app.id,
                app.organization_name,
                app.acronym or '',
                f"{app.applicant.firstname} {app.applicant.lastname}" if app.applicant else '',
                app.organization_email,
                app.organization_phone,
                app.address,
                app.status.value,
                app.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                app.last_modified.strftime('%Y-%m-%d %H:%M:%S'),
                app.certificate_number or '',
                app.cluster_information.cluster_of_intervention if app.cluster_information else ''
            ]
        else:
            row = [
                app.id,
                app.organization_name,
                f"{app.applicant.firstname} {app.applicant.lastname}" if app.applicant else '',
                app.status.value,
                app.submitted_at.strftime('%Y-%m-%d')
            ]
        writer.writerow(row)
    
    # Create response
    output.seek(0)
    filename = f"applications_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
    
    return {
        'downloadUrl': f'data:text/csv;charset=utf-8,{output.getvalue()}',
        'filename': filename
    }

def generate_excel_report(applications, report_type, start_date, end_date):
    """Generate Excel report"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Applications Report"
    
    # Headers
    if report_type == 'detailed':
        headers = [
            'Application ID', 'Organization Name', 'Acronym', 'Applicant Name', 
            'Email', 'Phone', 'Address', 'Status', 'Submitted At',
            'Last Modified', 'Certificate Number', 'Cluster of Intervention'
        ]
    else:
        headers = [
            'Application ID', 'Organization Name', 'Applicant Name', 
            'Status', 'Submitted At'
        ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row, app in enumerate(applications, 2):
        if report_type == 'detailed':
            data = [
                app.id,
                app.organization_name,
                app.acronym or '',
                f"{app.applicant.firstname} {app.applicant.lastname}" if app.applicant else '',
                app.organization_email,
                app.organization_phone,
                app.address,
                app.status.value,
                app.submitted_at,
                app.last_modified,
                app.certificate_number or '',
                app.cluster_information.cluster_of_intervention if app.cluster_information else ''
            ]
        else:
            data = [
                app.id,
                app.organization_name,
                f"{app.applicant.firstname} {app.applicant.lastname}" if app.applicant else '',
                app.status.value,
                app.submitted_at
            ]
        
        for col, value in enumerate(data, 1):
            worksheet.cell(row=row, column=col, value=value)
    
    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    filename = f"applications_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    return {
        'downloadUrl': f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{output.getvalue()}',
        'filename': filename
    }

def generate_pdf_report(applications, report_type, start_date, end_date, admin):
    """Generate PDF report"""
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width/2, height-50, "Rwanda Governance Board")
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, height-70, "Applications Report")
    
    # Report info
    p.setFont("Helvetica", 10)
    p.drawString(50, height-100, f"Report Type: {report_type.title()}")
    p.drawString(50, height-115, f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    p.drawString(50, height-130, f"Generated by: {admin.firstname} {admin.lastname}")
    p.drawString(50, height-145, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    p.drawString(50, height-160, f"Total Applications: {len(applications)}")
    
    # Statistics
    y_pos = height - 190
    if applications:
        status_counts = {}
        
        for app in applications:
            status = app.status.value
            status_counts[status] = status_counts.get(status, 0) + 1
        
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y_pos, "Summary Statistics:")
        y_pos -= 20
        
        p.setFont("Helvetica", 10)
        p.drawString(50, y_pos, "By Status:")
        y_pos -= 15
        for status, count in status_counts.items():
            p.drawString(70, y_pos, f"• {status}: {count}")
            y_pos -= 12
    
    # Applications list (summary)
    y_pos -= 30
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y_pos, "Applications:")
    y_pos -= 20
    
    p.setFont("Helvetica", 8)
    headers = ["ID", "Organization", "Applicant", "Status", "Submitted"]
    col_widths = [40, 150, 120, 80, 80]
    x_positions = [50]
    for width in col_widths[:-1]:
        x_positions.append(x_positions[-1] + width)
    
    # Draw headers
    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y_pos, header)
    y_pos -= 15
    
    # Draw applications
    for app in applications[:30]:  # Limit to first 30 for space
        if y_pos < 100:  # Start new page if needed
            p.showPage()
            y_pos = height - 50
        
        data = [
            str(app.id),
            app.organization_name[:20] + "..." if len(app.organization_name) > 20 else app.organization_name,
            f"{app.applicant.firstname} {app.applicant.lastname}"[:15] if app.applicant else "",
            app.status.value[:10],
            app.submitted_at.strftime('%m/%d/%Y')
        ]
        
        for i, value in enumerate(data):
            p.drawString(x_positions[i], y_pos, str(value))
        y_pos -= 12
    
    if len(applications) > 30:
        p.drawString(50, y_pos-10, f"... and {len(applications) - 30} more applications")
    
    # Footer
    p.setFont("Helvetica", 8)
    p.drawString(50, 30, f"Page 1 - Generated by RGB Portal - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    filename = f"applications_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
    
    return {
        'downloadUrl': f'data:application/pdf;base64,{buffer.getvalue()}',
        'filename': filename
    }

@bp.route('/applications/<int:application_id>/assign', methods=['POST'])
@admin_required(roles=[AdminRole.DIVISION_MANAGER, AdminRole.HOD, AdminRole.SECRETARY_GENERAL, AdminRole.CEO])
def assign_application(application_id):
    """Assign application to specific admin"""
    try:
        current_admin = get_current_user()
        data = request.get_json()
        
        application = OrganizationApplication.query.get_or_404(application_id)
        admin_id = data.get('admin_id')
        
        if not admin_id:
            return jsonify({'error': 'Admin ID is required'}), 400
        
        target_admin = Admin.query.get_or_404(admin_id)
        
        # Check if assignment is valid based on current status and roles
        valid_assignments = {
            ApplicationStatus.PENDING: [AdminRole.FBO_OFFICER],
            ApplicationStatus.DRAFT: [AdminRole.DIVISION_MANAGER],
            ApplicationStatus.DM_REVIEW: [AdminRole.HOD],
            ApplicationStatus.HOD_REVIEW: [AdminRole.SECRETARY_GENERAL],
            ApplicationStatus.SG_REVIEW: [AdminRole.CEO]
        }
        
        if application.status in valid_assignments:
            if target_admin.role not in valid_assignments[application.status]:
                return jsonify({'error': 'Invalid assignment for current application status'}), 400
        
        # Assign application
        application.processed_by_id = admin_id
        application.last_modified = datetime.utcnow()
        
        # Create notification for assigned admin
        notification = Notification(
            admin_id=admin_id,
            application_id=application_id,
            type=NotificationType.STATUS_CHANGE,
            title='Application Assigned',
            message=f'Application for {application.organization_name} has been assigned to you for review'
        )
        db.session.add(notification)
        db.session.commit()
        
        return jsonify({
            'message': 'Application assigned successfully',
            'application': application.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to assign application: {str(e)}'}), 500

@bp.route('/system/settings', methods=['GET'])
@admin_required(roles=[AdminRole.CEO])
def get_system_settings():
    """Get system settings (CEO only)"""
    try:
        # This would typically fetch from a settings table
        # For now, return default settings
        settings = {
            'application_fee': 50000,  # RWF
            'max_file_size': 16 * 1024 * 1024,  # 16MB
            'allowed_file_types': ['pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png'],
            'auto_assignment': True,
            'email_notifications': True,
            'sms_notifications': False,
            'maintenance_mode': False,
            'registration_open': True
        }
        
        return jsonify({'settings': settings})
        
    except Exception as e:
        return jsonify({'error': f'Failed to get settings: {str(e)}'}), 500

@bp.route('/system/settings', methods=['PUT'])
@admin_required(roles=[AdminRole.CEO])
def update_system_settings():
    """Update system settings (CEO only)"""
    try:
        data = request.get_json()
        
        # Validate settings
        allowed_settings = [
            'application_fee', 'max_file_size', 'allowed_file_types',
            'auto_assignment', 'email_notifications', 'sms_notifications',
            'maintenance_mode', 'registration_open'
        ]
        
        # In a real implementation, you would save these to a settings table
        # For now, just validate and return success
        
        updated_settings = {}
        for key, value in data.items():
            if key in allowed_settings:
                updated_settings[key] = value
        
        return jsonify({
            'message': 'Settings updated successfully',
            'settings': updated_settings
        })
        
    except Exception as e:
        return jsonify({'error': f'Failed to update settings: {str(e)}'}), 500

@bp.route('/analytics/trends', methods=['GET'])
@admin_required(roles=[AdminRole.CEO, AdminRole.SECRETARY_GENERAL])
def get_analytics_trends():
    """Get detailed analytics and trends"""
    try:
        # Application trends over time
        trends = {}
        
        # Monthly application trends (last 12 months)
        monthly_data = []
        for i in range(12):
            date = datetime.now() - timedelta(days=30 * i)
            start_date = date.replace(day=1)
            if i == 0:
                end_date = datetime.now()
            else:
                end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            applications = OrganizationApplication.query.filter(
                and_(
                    OrganizationApplication.submitted_at >= start_date,
                    OrganizationApplication.submitted_at <= end_date
                )
            ).all()
            
            status_breakdown = {}
            
            for app in applications:
                status = app.status.value
                status_breakdown[status] = status_breakdown.get(status, 0) + 1
            
            monthly_data.append({
                'month': date.month,
                'year': date.year,
                'total': len(applications),
                'status_breakdown': status_breakdown
            })
        
        trends['monthly'] = list(reversed(monthly_data))
        
        # Processing time analysis
        completed_apps = OrganizationApplication.query.filter(
            OrganizationApplication.status.in_([ApplicationStatus.APPROVED, ApplicationStatus.REJECTED])
        ).all()
        
        processing_times = []
        for app in completed_apps:
            if app.submitted_at and app.last_modified:
                days = (app.last_modified - app.submitted_at).days
                processing_times.append({
                    'application_id': app.id,
                    'days': days,
                    'status': app.status.value
                })
        
        if processing_times:
            avg_processing_time = sum(item['days'] for item in processing_times) / len(processing_times)
            trends['avg_processing_time'] = round(avg_processing_time, 1)
            trends['processing_times'] = processing_times[-50:]  # Last 50 completed applications
        else:
            trends['avg_processing_time'] = 0
            trends['processing_times'] = []
        
        return jsonify({'trends': trends})
        
    except Exception as e:
        return jsonify({'error': f'Failed to get analytics: {str(e)}'}), 500

@bp.route('/audit/logs', methods=['GET'])
@admin_required(roles=[AdminRole.CEO, AdminRole.SECRETARY_GENERAL])
def get_audit_logs():
    """Get audit logs for system activities"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        # In a real implementation, you would have an audit log table
        # For now, return recent application changes as audit logs
        
        applications = OrganizationApplication.query.order_by(
            OrganizationApplication.last_modified.desc()
        ).paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        logs = []
        for app in applications.items:
            logs.append({
                'id': app.id,
                'action': f'Application {app.status.value}',
                'entity_type': 'Application',
                'entity_id': app.id,
                'entity_name': app.organization_name,
                'user': f"{app.processor.firstname} {app.processor.lastname}" if app.processor else "System",
                'timestamp': app.last_modified.isoformat(),
                'details': {
                    'status': app.status.value,
                    'comments': [comment.to_dict() for comment in app.comments]

                }
            })
        
        return jsonify({
            'logs': logs,
            'pagination': {
                'page': applications.page,
                'pages': applications.pages,
                'per_page': applications.per_page,
                'total': applications.total
            }
        })
        
    except Exception as e:
        print(e)
        return jsonify({'error': f'Failed to get audit logs: {str(e)}'}), 500